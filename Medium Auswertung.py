
# coding: utf-8

import numpy as np
import openpyxl as op
import csv
import os

#Datenabfrage
ms_data = input('Laufnummer? ')
medium = input('Medium? ')
datum2 = input('Datum CLSA? dd/mm/yyyy ')
alkan = input('Alkanraster? ')

#Erstellen der Excel-Liste
path_script = r"C:\Analysis"
wb = op.Workbook()
wb.create_sheet(medium)
Auswertung = wb[medium]

daten1 = (
    ['Laufnummer:', ms_data],
	['Datum CLSA:', datum2],
    ['Alkanraster:', alkan],
	['Medium:', medium]
)
row = 1
col = 7
for item1, item2 in (daten1):
    Auswertung.cell(row=row, column=col, value=item1)
    Auswertung.cell(row=row, column=col+1, value=item2)
    row += 1
    

#Report Daten kopieren
path_report = os.path.join(path_script, "Report.tsv")
report = csv.reader(open(path_report), delimiter='\t')
a = 1
for i in report:
    Auswertung.cell(row=a, column=1, value=i[3])
    Auswertung.cell(row=a, column=2, value=i[4])
    a += 1
    
Auswertung.cell(row=1, column=1, value='RT [min]')
Auswertung.cell(row=1, column=2, value='RI')
Auswertung.cell(row=1, column=3, value='Verbindung')

#Formatierung
from openpyxl.styles import Font
font_bold = Font(name='Calibri',
               size=11,
               bold=True,
               italic=False,
                                )
 
Auswertung['A1'].font=font_bold
Auswertung['B1'].font=font_bold
Auswertung['C1'].font=font_bold
 
list1 = ['G'+str(x) for x in range (1,5)]
for i in list1:
    Auswertung[i].font=font_bold
list2 = ['I'+str(x) for x in range (1,5)]
for i in list2:
    Auswertung[i].font=font_bold

Auswertung.column_dimensions["A"].width = 12
Auswertung.column_dimensions["B"].width = 12
Auswertung.column_dimensions["C"].width = 50
Auswertung.column_dimensions["G"].width = 18
Auswertung.column_dimensions["H"].width = 18
Auswertung.column_dimensions["I"].width = 18
Auswertung.column_dimensions["J"].width = 18

from openpyxl.styles.borders import Border, Side
 
thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))
 
report = csv.reader(open(path_report), delimiter='\t')
row_count = sum(1 for row in report)
 
list3 = ['A'+str(x) for x in range (1, row_count+1)]
for i in list3:
    Auswertung[i].border=thin_border

list4 = ['B'+str(x) for x in range (1, row_count+1)]
for i in list4:
    Auswertung[i].border=thin_border

list5 = ['C'+str(x) for x in range (1, row_count+1)]
for i in list5:
    Auswertung[i].border=thin_border

#Erstellen der Excel-Liste
path_blank = os.path.join(path_script, medium+" Blank.xlsx")
wb.save(path_blank)


